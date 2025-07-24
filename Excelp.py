import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, Reference
import matplotlib.pyplot as plt
from tkinter import Tk, filedialog, messagebox, Button, Label
import os

class ExcelReportGenerator:
    def __init__(self):
        self.df = None
        self.wb = None
        
    def load_csv(self, file_path):
        """Load CSV file into pandas DataFrame"""
        try:
            self.df = pd.read_csv(file_path)
            return True
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load CSV: {str(e)}")
            return False
    
    def create_pivot_tables(self):
        """Generate pivot tables from the loaded data"""
        pivot_tables = []
        
        # Example pivot table - customize based on your data
        if 'Category' in self.df.columns and 'Sales' in self.df.columns:
            pivot1 = pd.pivot_table(self.df, 
                                  values='Sales', 
                                  index='Category', 
                                  aggfunc=['sum', 'mean', 'count'])
            pivot_tables.append(('Sales by Category', pivot1))
        
        if 'Date' in self.df.columns and 'Sales' in self.df.columns:
            self.df['Date'] = pd.to_datetime(self.df['Date'])
            self.df['Month'] = self.df['Date'].dt.month_name()
            pivot2 = pd.pivot_table(self.df,
                                  values='Sales',
                                  index='Month',
                                  aggfunc='sum')
            pivot_tables.append(('Monthly Sales', pivot2))
        
        return pivot_tables
    
    def generate_charts(self, sheet, data_range, title):
        """Create and embed charts in the Excel sheet"""
        chart = BarChart()
        chart.title = title
        chart.style = 13
        chart.y_axis.title = 'Values'
        chart.x_axis.title = 'Categories'
        
        data = Reference(sheet, min_col=data_range[0], min_row=data_range[1],
                        max_col=data_range[2], max_row=data_range[3])
        cats = Reference(sheet, min_col=data_range[0]-1, min_row=data_range[1]+1,
                        max_row=data_range[3])
        
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        
        sheet.add_chart(chart, f"F{data_range[1]}")
        
    def apply_styling(self, sheet):
        """Apply professional styling to Excel sheet"""
        # Header styling
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        
        # Adjust column widths
        for column in sheet.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            sheet.column_dimensions[column[0].column_letter].width = adjusted_width
    
    def generate_summary_stats(self, sheet):
        """Add summary statistics section"""
        if self.df is None:
            return
            
        # Summary section header
        sheet.append([])
        summary_header = sheet.max_row + 1
        sheet.cell(row=summary_header, column=1, value="Summary Statistics").font = Font(bold=True, size=12)
        
        # Basic stats
        sheet.append(["Total Records:", len(self.df)])
        
        # Numeric columns stats
        numeric_cols = self.df.select_dtypes(include=['number']).columns
        for col in numeric_cols:
            sheet.append([f"{col} - Mean:", self.df[col].mean()])
            sheet.append([f"{col} - Median:", self.df[col].median()])
            sheet.append([f"{col} - Std Dev:", self.df[col].std()])
            sheet.append([f"{col} - Min:", self.df[col].min()])
            sheet.append([f"{col} - Max:", self.df[col].max()])
            sheet.append([])  # Add empty row between metrics
    
    def generate_report(self, output_path):
        """Generate the complete Excel report"""
        try:
            self.wb = openpyxl.Workbook()
            
            # Create data sheet
            data_sheet = self.wb.active
            data_sheet.title = "Raw Data"
            for r in dataframe_to_rows(self.df, index=False, header=True):
                data_sheet.append(r)
            self.apply_styling(data_sheet)
            
            # Create analysis sheet
            analysis_sheet = self.wb.create_sheet("Analysis")
            pivot_tables = self.create_pivot_tables()
            
            for name, pivot in pivot_tables:
                # Add pivot table title
                title_row = analysis_sheet.max_row + 2
                analysis_sheet.cell(row=title_row, column=1, value=name).font = Font(bold=True, size=12)
                
                # Add pivot table data
                for r in dataframe_to_rows(pivot, index=True, header=True):
                    analysis_sheet.append(r)
                
                # Generate chart for the pivot table
                data_start = title_row + 2
                data_end = data_start + len(pivot)
                self.generate_charts(
                    analysis_sheet,
                    (2, data_start, len(pivot.columns)+1, data_end),
                    name
                )
            
            self.apply_styling(analysis_sheet)
            self.generate_summary_stats(analysis_sheet)
            
            # Save the workbook
            self.wb.save(output_path)
            return True
        except Exception as e:
            messagebox.showerror("Error", f"Report generation failed: {str(e)}")
            return False


class ReportGeneratorGUI:
    def __init__(self):
        self.root = Tk()
        self.root.title("Excel Report Generator")
        self.generator = ExcelReportGenerator()
        
        # Configure window size and position
        self.root.geometry("400x300")
        self.root.resizable(False, False)
        
        # Create UI elements
        self.create_widgets()
        
    def create_widgets(self):
        """Create the GUI interface"""
        # Header
        Label(self.root, text="Excel Report Generator", font=("Arial", 16, "bold")).pack(pady=20)
        
        # Instructions
        Label(self.root, text="1. Load your CSV file\n2. Generate Excel report", font=("Arial", 10)).pack(pady=10)
        
        # Buttons frame
        button_frame = Label(self.root)
        button_frame.pack(pady=20)
        
        # Action buttons
        Button(button_frame, text="Load CSV File", command=self.load_file, width=15, height=2).grid(row=0, column=0, padx=10)
        Button(button_frame, text="Generate Report", command=self.generate_report, width=15, height=2).grid(row=0, column=1, padx=10)
        
        # Exit button
        Button(self.root, text="Exit", command=self.root.quit, width=10).pack(pady=10)
    
    def load_file(self):
        """Handle file loading"""
        file_path = filedialog.askopenfilename(
            title="Select CSV File",
            filetypes=[("CSV Files", "*.csv"), ("All Files", "*.*")]
        )
        if file_path:
            if self.generator.load_csv(file_path):
                messagebox.showinfo("Success", "CSV file loaded successfully!")
    
    def generate_report(self):
        """Handle report generation"""
        if self.generator.df is None:
            messagebox.showwarning("Warning", "Please load a CSV file first")
            return
            
        output_path = filedialog.asksaveasfilename(
            title="Save Excel Report",
            defaultextension=".xlsx",
            filetypes=[("Excel Files", "*.xlsx")],
            initialfile="report.xlsx"
        )
        
        if output_path:
            if self.generator.generate_report(output_path):
                messagebox.showinfo(
                    "Success", 
                    f"Report generated successfully!\n\nSaved to:\n{output_path}"
                )
                # Optionally open the file
                if messagebox.askyesno("Open Report", "Would you like to open the report now?"):
                    try:
                        os.startfile(output_path)
                    except:
                        messagebox.showinfo(
                            "Info", 
                            "Could not open file automatically. Please open it manually."
                        )
    
    def run(self):
        """Run the application"""
        self.root.mainloop()


if __name__ == "__main__":
    # Create sample CSV if it doesn't exist
    sample_csv = "sample_sales_data.csv"
    if not os.path.exists(sample_csv):
        sample_data = """Date,Category,Product,Sales,Units
2023-01-01,Electronics,TV,1500,5
2023-01-01,Furniture,Sofa,1200,3
2023-01-02,Electronics,Phone,800,10
2023-01-02,Furniture,Table,400,8
2023-01-03,Electronics,Laptop,2000,4
2023-01-04,Furniture,Chair,300,12
2023-01-05,Electronics,Tablet,600,7
2023-01-06,Furniture,Cabinet,750,2
2023-02-01,Electronics,TV,1600,6
2023-02-02,Furniture,Sofa,1250,4
2023-02-03,Electronics,Phone,850,9
2023-02-04,Electronics,Laptop,2100,5"""
        
        with open(sample_csv, 'w') as f:
            f.write(sample_data)
    
    # Run the application
    app = ReportGeneratorGUI()
    app.run()